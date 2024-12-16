import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
from datetime import datetime

def process_excel_file():
    # 메인 윈도우 생성
    root = tk.Tk()
    root.title("엑셀 파일 처리")
    root.geometry("400x150")
    
    # 진행률 표시 레이블
    status_label = tk.Label(root, text="대기 중...")
    status_label.pack(pady=10)
    
    # 진행률 바 생성
    progress_bar = ttk.Progressbar(root, length=300, mode='determinate')
    progress_bar.pack(pady=10)
    
    # 파일 선택 다이얼로그
    input_file = filedialog.askopenfilename(
        title="엑셀 파일을 선택하세요",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    if not input_file:  # 파일 선택 취소시
        root.destroy()
        return
    
    try:
        # 엑셀 파일 읽기
        status_label.config(text="파일을 읽는 중...")
        root.update()
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
        
        # 헤더 찾기
        status_label.config(text="'검색량' 열을 찾는 중...")
        root.update()
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        search_col_idx = None
        for idx, header in enumerate(header_row, 1):
            if header == '검색량':
                search_col_idx = idx
                break
                
        if search_col_idx is None:
            messagebox.showerror("에러", "'검색량' 열을 찾을 수 없습니다.")
            root.destroy()
            return
            
        # 새 워크북 생성
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        
        # 헤더 복사
        status_label.config(text="헤더를 복사하는 중...")
        root.update()
        for col in range(1, sheet.max_column + 1):
            new_sheet.cell(1, col, sheet.cell(1, col).value)
        
        # 전체 행 수 계산
        total_rows = sheet.max_row - 1  # 헤더 제외
        
        # 데이터 필터링
        status_label.config(text="데이터 필터링 중...")
        root.update()
        new_row = 2
        filtered_count = 0
        
        for row in range(2, sheet.max_row + 1):
            # 진행률 업데이트
            progress = (row - 2) / total_rows * 100
            progress_bar['value'] = progress
            status_label.config(text=f"처리 중... ({row-1}/{total_rows} 행)")
            root.update()
            
            search_value = sheet.cell(row, search_col_idx).value
            try:
                # 숫자가 문자열로 되어있을 경우 쉼표 제거
                if isinstance(search_value, str):
                    search_value = int(search_value.replace(',', ''))
                else:
                    search_value = int(search_value)
                    
                if search_value >= 8000:
                    for col in range(1, sheet.max_column + 1):
                        new_sheet.cell(new_row, col, sheet.cell(row, col).value)
                    new_row += 1
                    filtered_count += 1
            except (ValueError, TypeError):
                continue
        
        # 저장
        status_label.config(text="파일 저장 중...")
        progress_bar['value'] = 100
        root.update()
        
        # 현재 날짜와 시간을 포함한 파일명 생성
        current_time = datetime.now().strftime('%Y-%m-%d_%H-%M')
        file_dir = os.path.dirname(input_file)
        file_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(file_dir, f"{file_name}_8000del_{current_time}.xlsx")
        new_wb.save(output_file)
        
        messagebox.showinfo("완료", f"처리가 완료되었습니다!\n\n- 전체 데이터: {total_rows}행\n- 필터링된 데이터: {filtered_count}행\n\n저장 경로:\n{output_file}")
            
    except Exception as e:
        messagebox.showerror("에러", f"처리 중 오류가 발생했습니다:\n{str(e)}")
    
    finally:
        root.destroy()

if __name__ == "__main__":
    process_excel_file()